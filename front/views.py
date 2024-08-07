from django.shortcuts import render, redirect
from datetime import datetime
from decimal import Decimal
from dataBridge import settings
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.files.storage import FileSystemStorage
from . import models
from . import constants
from django.db.models import Max, Min, Avg, F
from django.db.models.functions import TruncDate
import os
import math
import pathlib
import csv
import environ
import time
import json
import random
import openpyxl
from django.db import connections
from django.contrib.messages import get_messages
import re
from collections import defaultdict
# import pandas as pd
# from statsmodels.tsa.holtwinters import ExponentialSmoothing

import environ
env = environ.Env()
environ.Env.read_env()

context = {}
context['project_name'] = env("PROJECT_NAME")
context['client_name'] = env("CLIENT_NAME")

# Create your views here.


def removeSeconds(timestamp):
    # Split the timestamp into date and time parts
    date_part, time_part = timestamp.split('T')

    # Remove the seconds from the time part
    time_part_without_seconds = time_part[:5]

    # Recombine the date and time parts
    converted_timestamp = f"{date_part}T{time_part_without_seconds}"
    return converted_timestamp


def createMilisecondsByDate(date_string):
    # Define the time duration
    hours = -6
    minutes = 0

    # Calculate the total milliseconds
    timezone_milliseconds = (hours * 3600 * 1000) + (minutes * 60 * 1000)

    # Parse the datetime string manually
    if '+' in date_string:
        dt_object = datetime.strptime(date_string, '%Y-%m-%dT%H:%M:%S%z')
        # Access individual components
        year = dt_object.year
        month = dt_object.month
        day = dt_object.day
        hour = dt_object.hour
        minute = dt_object.minute
        second = dt_object.second
    else:
        year, month, day, hour, minute, second = map(int, date_string.split(
            'T')[0].split('-') + date_string.split('T')[1].split(':'))
    datetime_tuple = (year, month, day, hour, minute, second, 0, 0, 0)

    # Calculate the milliseconds since the Unix epoch
    milliseconds = int(time.mktime(datetime_tuple) *
                       1000) + timezone_milliseconds
    return milliseconds


def random_color_code():
    # Generate random values for red, green, and blue components
    r = random.randint(0, 255)
    g = random.randint(0, 255)
    b = random.randint(0, 255)

    # Convert the RGB values to a hexadecimal color code
    color_code = "#{:02X}{:02X}{:02X}".format(r, g, b)

    return color_code


# Download already existed Excel files
def downloadExcel(request, sensor_type):
    if sensor_type == 'strain':
        file_path = (settings.MEDIA_ROOT +
                     "sample/" + "Strain Data.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'attachment; filename=' + \
                    os.path.basename(file_path)
                return response
    elif sensor_type == 'tilt':
        file_path = (settings.MEDIA_ROOT + "sample/" + "Tilt Data.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'attachment; filename=' + \
                    os.path.basename(file_path)
                return response
    elif sensor_type == 'displacement':
        file_path = (settings.MEDIA_ROOT + "sample/" +
                     "Displacement Data.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'attachment; filename=' + \
                    os.path.basename(file_path)
                return response
    elif sensor_type == 'settlement':
        file_path = (settings.MEDIA_ROOT + "sample/" +
                     "Settlement Data.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'attachment; filename=' + \
                    os.path.basename(file_path)
                return response
    elif sensor_type == 'vibration':
        file_path = (settings.MEDIA_ROOT +
                     "sample/" + "Vibration Data.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'attachment; filename=' + \
                    os.path.basename(file_path)
                return response


# If columns for every sensor type already exists, then we are collecting it from these below updated files, otherwise collecting it from constants file
def get_constants(sensor_type):
    if sensor_type == 'strain':
        file_path = "templates/constants/strain_columns.txt"
        f = open(file_path)
        columns = f.read()
        if columns == "":
            columns = json.dumps(constants.strain_columns)
            f = open(file_path, "w+")
            f.write(columns)
        f.close()
        return json.loads(columns)
    if sensor_type == 'tilt':
        file_path = "templates/constants/tilt_columns.txt"
        f = open(file_path)
        columns = f.read()
        if columns == "":
            columns = json.dumps(constants.tilt_columns)
            f = open(file_path, "w+")
            f.write(columns)
        f.close()
        return json.loads(columns)
    if sensor_type == 'displacement':
        file_path = "templates/constants/displacement_columns.txt"
        f = open(file_path)
        columns = f.read()
        if columns == "":
            columns = json.dumps(constants.displacement_columns)
            f = open(file_path, "w+")
            f.write(columns)
        f.close()
        return json.loads(columns)
    if sensor_type == 'settlement':
        file_path = "templates/constants/settlement_columns.txt"
        f = open(file_path)
        columns = f.read()
        if columns == "":
            columns = json.dumps(constants.settlement_columns)
            f = open(file_path, "w+")
            f.write(columns)
        f.close()
        return json.loads(columns)
    if sensor_type == 'vibration':
        file_path = "templates/constants/vibration_columns.txt"
        f = open(file_path)
        columns = f.read()
        if columns == "":
            columns = json.dumps(constants.vibration_columns)
            f = open(file_path, "w+")
            f.write(columns)
        f.close()
        return json.loads(columns)


# Getting sensor names(columns which are updated on the time of importing) by sensor types
def getSensorsByTypes(request):
    if request.method == "POST":
        sensor_type = request.POST['sensor_type']
        sensor_names = get_constants(sensor_type)
        columns = {k: v for k, v in sensor_names.items(
        ) if not v.startswith('test_method_')}

        if sensor_type == 'strain':
            min_time = (models.StrainData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
            max_time = (models.StrainData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
        if sensor_type == 'tilt':
            min_time = (models.TiltData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
            max_time = (models.TiltData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
        if sensor_type == 'displacement':
            min_time = (models.DisplacementData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
            max_time = (models.DisplacementData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
        if sensor_type == 'settlement':
            min_time = (models.SettlementData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
            max_time = (models.SettlementData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
        if sensor_type == 'vibration':
            min_time = (models.VibrationData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
            max_time = (models.VibrationData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
        return JsonResponse({
            'code': 200,
            'status': "SUCCESS",
            'columns': columns,
            'min_time': removeSeconds(min_time),
            'max_time': removeSeconds(max_time)
        })
    else:
        return JsonResponse({
            'code': 501,
            'status': "ERROR",
            'message': "There should be ajax method"
        })


# Get the exact columns which are uploaded on the time of importing
def getSensorCounts(sensor_type):
    if sensor_type == 'strain':
        column_names = []
        file_path = "templates/constants/strain_columns.txt"
        f = open(file_path)
        columns_constants = f.read()
        f.close()
        columns_constants = json.loads(columns_constants)

        columns_fields = [
            field.name for field in models.StrainData._meta.get_fields()]
        for each in columns_fields:
            if each.startswith('test_method_') and (each in columns_constants.keys()) and columns_constants[each] != each:
                column_names.append(columns_constants[each])
        return column_names
    if sensor_type == 'tilt':
        column_names = []
        file_path = "templates/constants/tilt_columns.txt"
        f = open(file_path)
        columns_constants = f.read()
        f.close()
        columns_constants = json.loads(columns_constants)

        columns_fields = [
            field.name for field in models.TiltData._meta.get_fields()]
        for each in columns_fields:
            if each.startswith('test_method_') and (each in columns_constants.keys()) and columns_constants[each] != each:
                column_names.append(columns_constants[each])
        return column_names
    if sensor_type == 'displacement':
        column_names = []
        file_path = "templates/constants/displacement_columns.txt"
        f = open(file_path)
        columns_constants = f.read()
        f.close()
        columns_constants = json.loads(columns_constants)

        columns_fields = [
            field.name for field in models.DisplacementData._meta.get_fields()]
        for each in columns_fields:
            if each.startswith('test_method_') and (each in columns_constants.keys()) and columns_constants[each] != each:
                column_names.append(columns_constants[each])
        return column_names
    if sensor_type == 'settlement':
        column_names = []
        file_path = "templates/constants/settlement_columns.txt"
        f = open(file_path)
        columns_constants = f.read()
        f.close()
        columns_constants = json.loads(columns_constants)

        columns_fields = [
            field.name for field in models.SettlementData._meta.get_fields()]
        for each in columns_fields:
            if each.startswith('test_method_') and (each in columns_constants.keys()) and columns_constants[each] != each:
                column_names.append(columns_constants[each])
        return column_names
    if sensor_type == 'vibration':
        column_names = []
        file_path = "templates/constants/vibration_columns.txt"
        f = open(file_path)
        columns_constants = f.read()
        f.close()
        columns_constants = json.loads(columns_constants)

        columns_fields = [
            field.name for field in models.VibrationData._meta.get_fields()]
        for each in columns_fields:
            if each.startswith('test_method_') and (each in columns_constants.keys()) and columns_constants[each] != each:
                column_names.append(columns_constants[each])
        return column_names


# Showing Index page if all the sensor data are present otherwise importing form will be displayed
def index(request):
    strain_data = models.StrainData.objects.count()
    if strain_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Strain Data Is Present.')
        return redirect('importExcel')
    tilt_data = models.TiltData.objects.count()
    if tilt_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Tilt Data Is Present.')
        return redirect('importExcel')
    displacement_data = models.DisplacementData.objects.count()
    if displacement_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Displacement Data Is Present.')
        return redirect('importExcel')
    settlement_data = models.SettlementData.objects.count()
    if settlement_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Settlement Data Is Present.')
        return redirect('importExcel')
    vibration_data = models.VibrationData.objects.count()
    if vibration_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Vibration Data Is Present.')
        return redirect('importExcel')
    # strain_sensor_counts = getSensorCounts('strain')
    # tilt_sensor_counts = getSensorCounts('tilt')
    # displacement_sensor_counts = getSensorCounts('displacement')
    # settlement_sensor_counts = getSensorCounts('settlement')
    # vibration_sensor_counts = getSensorCounts('vibration')
    # min_time = (models.StrainData.objects.all().order_by('date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
    # max_time = (models.StrainData.objects.all().order_by('-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S")
    sensor_types = {
        'keys': list(constants.sensor_types.keys()),
        'values': list(constants.sensor_types.values())
    }
    # context.update({'sensor_types':sensor_types, 'strain_sensor_counts': strain_sensor_counts, 'tilt_sensor_counts': tilt_sensor_counts, 'displacement_sensor_counts': displacement_sensor_counts, 'settlement_sensor_counts': settlement_sensor_counts, 'vibration_sensor_counts': vibration_sensor_counts, 'min_time': min_time, 'max_time': max_time})
    context.update({'sensor_types': sensor_types})
    return render(request, 'front/index.html', context)


# Importing Sensor Types and uploading it into the database
# and updating the column values by removing values in constants file
def importExcel(request):
    if request.method == "POST":
        if 'strain_required' in request.POST.keys():
            models.StrainData.objects.all().delete()
            f = open("templates/constants/strain_columns.txt", "w+")
            f.write("")
            f.close()
            if request.FILES.get('strain_gauge', None):
                file = request.FILES['strain_gauge']
                if os.path.splitext(str(file))[1] == '.xlsx' or os.path.splitext(str(file))[1] == '.xls':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "excel/", settings.MEDIA_ROOT + "/excel/")
                    fs.save(tmpname, file)
                    file_name = "excel/" + tmpname

                    wb = openpyxl.load_workbook(
                        settings.MEDIA_ROOT + file_name, data_only=True)
                    ws = wb.active

                    # Extract and clean column names from the Excel file (assuming they are in the first row)
                    # column_names = [re.sub(r'[^a-zA-Z0-9_]', '_', str(cell.value)) for cell in ws[1]]
                    column_names = [str(cell.value) for cell in ws[1]]

                    # Removing Date Time Column
                    del column_names[0]

                    # strain_columns = get_constants('strain')
                    strain_columns = constants.strain_columns
                    new_strain_columns = {}
                    for index, column in enumerate(column_names):
                        new_strain_columns[list(strain_columns.keys())[
                            index]] = column

                    f = open("templates/constants/strain_columns.txt", "w+")
                    columns = json.dumps(new_strain_columns)
                    f.write(columns)
                    f.close()

                    data_list = []

                    # Initialize a flag to skip the first row
                    skip_first_row = True

                    for row in ws.iter_rows(values_only=True):
                        if skip_first_row:
                            skip_first_row = False
                            continue  # Skip the first row
                        if not row[0]:
                            break
                        if len(data_list) > 1000:
                            models.StrainData.objects.bulk_create(data_list)
                            data_list = []
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(
                                row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None, test_method_26=row[26] if len(row) > 26 else None, test_method_27=row[27] if len(row) > 27 else None, test_method_28=row[28] if len(row) > 28 else None, test_method_29=row[29] if len(row) > 29 else None, test_method_30=row[30] if len(row) > 30 else None, test_method_31=row[31] if len(row) > 31 else None, test_method_32=row[32] if len(row) > 32 else None, test_method_33=row[33] if len(row) > 33 else None, test_method_34=row[34] if len(row) > 34 else None, test_method_35=row[35] if len(row) > 35 else None, test_method_36=row[36] if len(row) > 36 else None, test_method_37=row[37] if len(row) > 37 else None, test_method_38=row[38] if len(row) > 38 else None, test_method_39=row[39] if len(row) > 39 else None, test_method_40=row[40] if len(row) > 40 else None, test_method_41=row[41] if len(row) > 41 else None, test_method_42=row[42] if len(row) > 42 else None))
                        else:
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(
                                row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None, test_method_26=row[26] if len(row) > 26 else None, test_method_27=row[27] if len(row) > 27 else None, test_method_28=row[28] if len(row) > 28 else None, test_method_29=row[29] if len(row) > 29 else None, test_method_30=row[30] if len(row) > 30 else None, test_method_31=row[31] if len(row) > 31 else None, test_method_32=row[32] if len(row) > 32 else None, test_method_33=row[33] if len(row) > 33 else None, test_method_34=row[34] if len(row) > 34 else None, test_method_35=row[35] if len(row) > 35 else None, test_method_36=row[36] if len(row) > 36 else None, test_method_37=row[37] if len(row) > 37 else None, test_method_38=row[38] if len(row) > 38 else None, test_method_39=row[39] if len(row) > 39 else None, test_method_40=row[40] if len(row) > 40 else None, test_method_41=row[41] if len(row) > 41 else None, test_method_42=row[42] if len(row) > 42 else None))
                    models.StrainData.objects.bulk_create(data_list)
                    os.remove(settings.MEDIA_ROOT + file_name)
                elif os.path.splitext(str(file))[1] == '.csv':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "csv/", settings.MEDIA_ROOT + "/csv/")
                    fs.save(tmpname, file)
                    file_name = "csv/" + tmpname

                    with open(settings.MEDIA_ROOT + file_name, newline='', mode='r', encoding='ISO-8859-1') as csvfile:
                        reader = csv.DictReader(csvfile)
                        reader = csv.reader(csvfile)
                        first_row = next(reader, None)

                        # Removing Date Time Column
                        del first_row[0]

                        # strain_columns = get_constants('strain')
                        strain_columns = constants.strain_columns
                        new_strain_columns = {}
                        for index, column in enumerate(first_row):
                            new_strain_columns[list(strain_columns.keys())[
                                index]] = column

                        f = open("templates/constants/strain_columns.txt", "w+")
                        columns = json.dumps(new_strain_columns)
                        f.write(columns)
                        f.close()

                        next(reader)
                        data_list = []
                        for row in reader:
                            if len(data_list) > 1000:
                                models.StrainData.objects.bulk_create(
                                    data_list)
                                data_list = []
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(
                                    row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None, test_method_26=row[26] if len(row) > 26 else None, test_method_27=row[27] if len(row) > 27 else None, test_method_28=row[28] if len(row) > 28 else None, test_method_29=row[29] if len(row) > 29 else None, test_method_30=row[30] if len(row) > 30 else None, test_method_31=row[31] if len(row) > 31 else None, test_method_32=row[32] if len(row) > 32 else None, test_method_33=row[33] if len(row) > 33 else None, test_method_34=row[34] if len(row) > 34 else None, test_method_35=row[35] if len(row) > 35 else None, test_method_36=row[36] if len(row) > 36 else None, test_method_37=row[37] if len(row) > 37 else None, test_method_38=row[38] if len(row) > 38 else None, test_method_39=row[39] if len(row) > 39 else None, test_method_40=row[40] if len(row) > 40 else None, test_method_41=row[41] if len(row) > 41 else None, test_method_42=row[42] if len(row) > 42 else None))
                            else:
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(
                                    row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None, test_method_26=row[26] if len(row) > 26 else None, test_method_27=row[27] if len(row) > 27 else None, test_method_28=row[28] if len(row) > 28 else None, test_method_29=row[29] if len(row) > 29 else None, test_method_30=row[30] if len(row) > 30 else None, test_method_31=row[31] if len(row) > 31 else None, test_method_32=row[32] if len(row) > 32 else None, test_method_33=row[33] if len(row) > 33 else None, test_method_34=row[34] if len(row) > 34 else None, test_method_35=row[35] if len(row) > 35 else None, test_method_36=row[36] if len(row) > 36 else None, test_method_37=row[37] if len(row) > 37 else None, test_method_38=row[38] if len(row) > 38 else None, test_method_39=row[39] if len(row) > 39 else None, test_method_40=row[40] if len(row) > 40 else None, test_method_41=row[41] if len(row) > 41 else None, test_method_42=row[42] if len(row) > 42 else None))
                        models.StrainData.objects.bulk_create(data_list)
                        csvfile.close()
                        os.remove(settings.MEDIA_ROOT + file_name)
        if 'tilt_required' in request.POST.keys():
            models.TiltData.objects.all().delete()
            f = open("templates/constants/tilt_columns.txt", "w+")
            f.write("")
            f.close()
            if request.FILES.get('tilt', None):
                file = request.FILES['tilt']
                if os.path.splitext(str(file))[1] == '.xlsx' or os.path.splitext(str(file))[1] == '.xls':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "excel/", settings.MEDIA_ROOT + "/excel/")
                    fs.save(tmpname, file)
                    file_name = "excel/" + tmpname

                    wb = openpyxl.load_workbook(
                        settings.MEDIA_ROOT + file_name, data_only=True)
                    ws = wb.active

                    # Extract and clean column names from the Excel file (assuming they are in the first row)
                    # column_names = [re.sub(r'[^a-zA-Z0-9_]', '_', str(cell.value)) for cell in ws[1]]
                    column_names = [str(cell.value) for cell in ws[1]]

                    # Removing Date Time Column
                    del column_names[0]

                    # tilt_columns = get_constants('tilt')
                    tilt_columns = constants.tilt_columns
                    new_tilt_columns = {}
                    for index, column in enumerate(column_names):
                        new_tilt_columns[list(tilt_columns.keys())[
                            index]] = column

                    f = open("templates/constants/tilt_columns.txt", "w+")
                    columns = json.dumps(new_tilt_columns)
                    f.write(columns)
                    f.close()

                    data_list = []

                    # Initialize a flag to skip the first row
                    skip_first_row = True

                    for row in ws.iter_rows(values_only=True):
                        if skip_first_row:
                            skip_first_row = False
                            continue  # Skip the first row
                        if not row[0]:
                            break
                        if len(data_list) > 1000:
                            models.TiltData.objects.bulk_create(data_list)
                            data_list = []
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        else:
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                    models.TiltData.objects.bulk_create(data_list)
                    os.remove(settings.MEDIA_ROOT + file_name)
                elif os.path.splitext(str(file))[1] == '.csv':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "csv/", settings.MEDIA_ROOT + "/csv/")
                    fs.save(tmpname, file)
                    file_name = "csv/" + tmpname

                    with open(settings.MEDIA_ROOT + file_name, newline='', mode='r', encoding='ISO-8859-1') as csvfile:
                        reader = csv.DictReader(csvfile)
                        reader = csv.reader(csvfile)
                        first_row = next(reader, None)

                        # Removing Date Time Column
                        del first_row[0]

                        # tilt_columns = get_constants('tilt')
                        tilt_columns = constants.tilt_columns
                        new_tilt_columns = {}
                        for index, column in enumerate(first_row):
                            new_tilt_columns[list(tilt_columns.keys())[
                                index]] = column

                        f = open("templates/constants/tilt_columns.txt", "w+")
                        columns = json.dumps(new_tilt_columns)
                        f.write(columns)
                        f.close()

                        next(reader)
                        data_list = []
                        for row in reader:
                            if len(data_list) > 1000:
                                models.TiltData.objects.bulk_create(data_list)
                                data_list = []
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                            else:
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        models.TiltData.objects.bulk_create(data_list)
                        csvfile.close()
                        os.remove(settings.MEDIA_ROOT + file_name)
        if 'displacement_required' in request.POST.keys():
            models.DisplacementData.objects.all().delete()
            f = open("templates/constants/displacement_columns.txt", "w+")
            f.write("")
            f.close()
            if request.FILES.get('displacement', None):
                file = request.FILES['displacement']
                if os.path.splitext(str(file))[1] == '.xlsx' or os.path.splitext(str(file))[1] == '.xls':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "excel/", settings.MEDIA_ROOT + "/excel/")
                    fs.save(tmpname, file)
                    file_name = "excel/" + tmpname

                    wb = openpyxl.load_workbook(
                        settings.MEDIA_ROOT + file_name, data_only=True)
                    ws = wb.active

                    # Extract and clean column names from the Excel file (assuming they are in the first row)
                    # column_names = [re.sub(r'[^a-zA-Z0-9_]', '_', str(cell.value)) for cell in ws[1]]
                    column_names = [str(cell.value) for cell in ws[1]]

                    # Removing Date Time Column
                    del column_names[0]

                    # displacement_columns = get_constants('displacement')
                    displacement_columns = constants.displacement_columns
                    new_displacement_columns = {}
                    for index, column in enumerate(column_names):
                        new_displacement_columns[list(displacement_columns.keys())[
                            index]] = column

                    f = open("templates/constants/displacement_columns.txt", "w+")
                    columns = json.dumps(new_displacement_columns)
                    f.write(columns)
                    f.close()

                    data_list = []

                    # Initialize a flag to skip the first row
                    skip_first_row = True

                    for row in ws.iter_rows(values_only=True):
                        if skip_first_row:
                            skip_first_row = False
                            continue  # Skip the first row
                        if not row[0]:
                            break
                        if len(data_list) > 1000:
                            models.DisplacementData.objects.bulk_create(
                                data_list)
                            data_list = []
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        else:
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                    models.DisplacementData.objects.bulk_create(data_list)
                    os.remove(settings.MEDIA_ROOT + file_name)
                elif os.path.splitext(str(file))[1] == '.csv':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "csv/", settings.MEDIA_ROOT + "/csv/")
                    fs.save(tmpname, file)
                    file_name = "csv/" + tmpname

                    with open(settings.MEDIA_ROOT + file_name, newline='', mode='r', encoding='ISO-8859-1') as csvfile:
                        reader = csv.DictReader(csvfile)
                        reader = csv.reader(csvfile)
                        first_row = next(reader, None)

                        # Removing Date Time Column
                        del first_row[0]

                        # displacement_columns = get_constants('displacement')
                        displacement_columns = constants.displacement_columns
                        new_displacement_columns = {}
                        for index, column in enumerate(first_row):
                            new_displacement_columns[list(displacement_columns.keys())[
                                index]] = column

                        f = open(
                            "templates/constants/displacement_columns.txt", "w+")
                        columns = json.dumps(new_displacement_columns)
                        f.write(columns)
                        f.close()

                        next(reader)
                        data_list = []
                        for row in reader:
                            if len(data_list) > 1000:
                                models.DisplacementData.objects.bulk_create(
                                    data_list)
                                data_list = []
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                            else:
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        models.DisplacementData.objects.bulk_create(data_list)
                        csvfile.close()
                        os.remove(settings.MEDIA_ROOT + file_name)
        if 'settlement_required' in request.POST.keys():
            models.SettlementData.objects.all().delete()
            f = open("templates/constants/settlement_columns.txt", "w+")
            f.write("")
            f.close()
            if request.FILES.get('settlement', None):
                file = request.FILES['settlement']
                if os.path.splitext(str(file))[1] == '.xlsx' or os.path.splitext(str(file))[1] == '.xls':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "excel/", settings.MEDIA_ROOT + "/excel/")
                    fs.save(tmpname, file)
                    file_name = "excel/" + tmpname

                    wb = openpyxl.load_workbook(
                        settings.MEDIA_ROOT + file_name, data_only=True)
                    ws = wb.active

                    # Extract and clean column names from the Excel file (assuming they are in the first row)
                    # column_names = [re.sub(r'[^a-zA-Z0-9_]', '_', str(cell.value)) for cell in ws[1]]
                    column_names = [str(cell.value) for cell in ws[1]]

                    # Removing Date Time Column
                    del column_names[0]

                    # settlement_columns = get_constants('settlement')
                    settlement_columns = constants.settlement_columns
                    new_settlement_columns = {}
                    for index, column in enumerate(column_names):
                        new_settlement_columns[list(settlement_columns.keys())[
                            index]] = column

                    f = open("templates/constants/settlement_columns.txt", "w+")
                    columns = json.dumps(new_settlement_columns)
                    f.write(columns)
                    f.close()

                    data_list = []

                    # Initialize a flag to skip the first row
                    skip_first_row = True

                    for row in ws.iter_rows(values_only=True):
                        if skip_first_row:
                            skip_first_row = False
                            continue  # Skip the first row
                        if not row[0]:
                            break
                        if len(data_list) > 1000:
                            models.SettlementData.objects.bulk_create(
                                data_list)
                            data_list = []
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        else:
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                    models.SettlementData.objects.bulk_create(data_list)
                    os.remove(settings.MEDIA_ROOT + file_name)
                elif os.path.splitext(str(file))[1] == '.csv':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "csv/", settings.MEDIA_ROOT + "/csv/")
                    fs.save(tmpname, file)
                    file_name = "csv/" + tmpname

                    with open(settings.MEDIA_ROOT + file_name, newline='', mode='r', encoding='ISO-8859-1') as csvfile:
                        reader = csv.DictReader(csvfile)
                        reader = csv.reader(csvfile)
                        first_row = next(reader, None)

                        # Removing Date Time Column
                        del first_row[0]

                        # settlement_columns = get_constants('settlement')
                        settlement_columns = constants.settlement_columns
                        new_settlement_columns = {}
                        for index, column in enumerate(first_row):
                            new_settlement_columns[list(settlement_columns.keys())[
                                index]] = column

                        f = open(
                            "templates/constants/settlement_columns.txt", "w+")
                        columns = json.dumps(new_settlement_columns)
                        f.write(columns)
                        f.close()

                        next(reader)
                        data_list = []
                        for row in reader:
                            if len(data_list) > 1000:
                                models.SettlementData.objects.bulk_create(
                                    data_list)
                                data_list = []
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                            else:
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        models.SettlementData.objects.bulk_create(data_list)
                        csvfile.close()
                        os.remove(settings.MEDIA_ROOT + file_name)
        if 'vibration_required' in request.POST.keys():
            models.VibrationData.objects.all().delete()
            f = open("templates/constants/vibration_columns.txt", "w+")
            f.write("")
            f.close()
            if request.FILES.get('vibration', None):
                file = request.FILES['vibration']
                if os.path.splitext(str(file))[1] == '.xlsx' or os.path.splitext(str(file))[1] == '.xls':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "excel/", settings.MEDIA_ROOT + "/excel/")
                    fs.save(tmpname, file)
                    file_name = "excel/" + tmpname

                    wb = openpyxl.load_workbook(
                        settings.MEDIA_ROOT + file_name, data_only=True)
                    ws = wb.active

                    # Extract and clean column names from the Excel file (assuming they are in the first row)
                    # column_names = [re.sub(r'[^a-zA-Z0-9_]', '_', str(cell.value)) for cell in ws[1]]
                    column_names = [str(cell.value) for cell in ws[1]]

                    # Removing Date Time Column
                    del column_names[0]

                    # vibration_columns = get_constants('vibration')
                    vibration_columns = constants.vibration_columns
                    new_vibration_columns = {}
                    for index, column in enumerate(column_names):
                        new_vibration_columns[list(vibration_columns.keys())[
                            index]] = column

                    f = open("templates/constants/vibration_columns.txt", "w+")
                    columns = json.dumps(new_vibration_columns)
                    f.write(columns)
                    f.close()

                    data_list = []

                    # Initialize a flag to skip the first row
                    skip_first_row = True

                    for row in ws.iter_rows(values_only=True):
                        if skip_first_row:
                            skip_first_row = False
                            continue  # Skip the first row
                        if not row[0]:
                            break
                        if len(data_list) > 1000:
                            models.VibrationData.objects.bulk_create(data_list)
                            data_list = []
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        else:
                            data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                    models.VibrationData.objects.bulk_create(data_list)
                    os.remove(settings.MEDIA_ROOT + file_name)
                elif os.path.splitext(str(file))[1] == '.csv':
                    tmpname = str(datetime.now().microsecond) + \
                        os.path.splitext(str(file))[1]
                    fs = FileSystemStorage(
                        settings.MEDIA_ROOT + "csv/", settings.MEDIA_ROOT + "/csv/")
                    fs.save(tmpname, file)
                    file_name = "csv/" + tmpname

                    with open(settings.MEDIA_ROOT + file_name, newline='', mode='r', encoding='ISO-8859-1') as csvfile:
                        reader = csv.DictReader(csvfile)
                        reader = csv.reader(csvfile)
                        first_row = next(reader, None)

                        # Removing Date Time Column
                        del first_row[0]

                        # vibration_columns = get_constants('vibration')
                        vibration_columns = constants.vibration_columns
                        new_vibration_columns = {}
                        for index, column in enumerate(first_row):
                            new_vibration_columns[list(vibration_columns.keys())[
                                index]] = column

                        f = open(
                            "templates/constants/vibration_columns.txt", "w+")
                        columns = json.dumps(new_vibration_columns)
                        f.write(columns)
                        f.close()

                        next(reader)
                        data_list = []
                        for row in reader:
                            if len(data_list) > 1000:
                                models.VibrationData.objects.bulk_create(
                                    data_list)
                                data_list = []
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                            else:
                                data_list.append(models.StrainData(date_time=datetime.strptime(row[0], "%d-%m-%Y %H:%M:%S") if len(row[0].split(":")) > 2 else datetime.strptime(row[0], "%d-%m-%Y %H:%M"), test_method_1=row[1] if len(row) > 1 else None, test_method_2=row[2] if len(row) > 2 else None, test_method_3=row[3] if len(row) > 3 else None, test_method_4=row[4] if len(row) > 4 else None, test_method_5=row[5] if len(row) > 5 else None, test_method_6=row[6] if len(row) > 6 else None, test_method_7=row[7] if len(row) > 7 else None, test_method_8=row[8] if len(row) > 8 else None, test_method_9=row[9] if len(row) > 9 else None, test_method_10=row[10] if len(row) > 10 else None, test_method_11=row[11] if len(row) > 11 else None, test_method_12=row[12] if len(
                                    row) > 12 else None, test_method_13=row[13] if len(row) > 13 else None, test_method_14=row[14] if len(row) > 14 else None, test_method_15=row[15] if len(row) > 15 else None, test_method_16=row[16] if len(row) > 16 else None, test_method_17=row[17] if len(row) > 17 else None, test_method_18=row[18] if len(row) > 18 else None, test_method_19=row[19] if len(row) > 19 else None, test_method_20=row[20] if len(row) > 20 else None, test_method_21=row[21] if len(row) > 21 else None, test_method_22=row[22] if len(row) > 22 else None, test_method_23=row[23] if len(row) > 23 else None, test_method_24=row[24] if len(row) > 24 else None, test_method_25=row[25] if len(row) > 25 else None))
                        models.VibrationData.objects.bulk_create(data_list)
                        csvfile.close()
                        os.remove(settings.MEDIA_ROOT + file_name)
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.success(request, 'Data Uploaded Successfully.')
        return redirect('index')
    return render(request, 'front/import.html', context)


def find_key_by_value(dictionary, value_to_find):
    for key, value in dictionary.items():
        if value == value_to_find:
            return key
    # If the value is not found, you can return None or raise an exception.
    return None  # or raise ValueError("Value not found")


# Generating single type chart data by highcharts' recommendation
def getChartData(request):
    if request.method == "POST":
        chart_type = request.POST['chart_type']
        sensor_type = request.POST['sensor_type']
        from_time = request.POST['from_time']
        to_time = request.POST['to_time']
        from_miliseconds = int(datetime.fromisoformat(
            from_time).timestamp() * 1000)
        to_miliseconds = int(datetime.fromisoformat(
            to_time).timestamp() * 1000)
        if from_miliseconds > to_miliseconds:
            return JsonResponse({
                'code': 505,
                'status': "ERROR",
                'message': "From time should not exceeds To time "
            })
        if chart_type == 'max':
            if sensor_type == 'strain':
                data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25'),
                    test_method_26=Max('test_method_26'),
                    test_method_27=Max('test_method_27'),
                    test_method_28=Max('test_method_28'),
                    test_method_29=Max('test_method_29'),
                    test_method_30=Max('test_method_30'),
                    test_method_31=Max('test_method_31'),
                    test_method_32=Max('test_method_32'),
                    test_method_33=Max('test_method_33'),
                    test_method_34=Max('test_method_34'),
                    test_method_35=Max('test_method_35'),
                    test_method_36=Max('test_method_36'),
                    test_method_37=Max('test_method_37'),
                    test_method_38=Max('test_method_38'),
                    test_method_39=Max('test_method_39'),
                    test_method_40=Max('test_method_40'),
                    test_method_41=Max('test_method_41'),
                    test_method_42=Max('test_method_42')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'tilt':
                data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'displacement':
                data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'settlement':
                data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'vibration':
                data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            return JsonResponse({
                'code': 200,
                'status': "SUCCESS",
                'result': {'series': series},
            })
        elif chart_type == 'min':
            if sensor_type == 'strain':
                data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25'),
                    test_method_26=Min('test_method_26'),
                    test_method_27=Min('test_method_27'),
                    test_method_28=Min('test_method_28'),
                    test_method_29=Min('test_method_29'),
                    test_method_30=Min('test_method_30'),
                    test_method_31=Min('test_method_31'),
                    test_method_32=Min('test_method_32'),
                    test_method_33=Min('test_method_33'),
                    test_method_34=Min('test_method_34'),
                    test_method_35=Min('test_method_35'),
                    test_method_36=Min('test_method_36'),
                    test_method_37=Min('test_method_37'),
                    test_method_38=Min('test_method_38'),
                    test_method_39=Min('test_method_39'),
                    test_method_40=Min('test_method_40'),
                    test_method_41=Min('test_method_41'),
                    test_method_42=Min('test_method_42')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'tilt':
                data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'displacement':
                data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'settlement':
                data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'vibration':
                data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            return JsonResponse({
                'code': 200,
                'status': "SUCCESS",
                'result': {'series': series},
            })
        elif chart_type == 'avg':
            if sensor_type == 'strain':
                data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25'),
                    test_method_26=Avg('test_method_26'),
                    test_method_27=Avg('test_method_27'),
                    test_method_28=Avg('test_method_28'),
                    test_method_29=Avg('test_method_29'),
                    test_method_30=Avg('test_method_30'),
                    test_method_31=Avg('test_method_31'),
                    test_method_32=Avg('test_method_32'),
                    test_method_33=Avg('test_method_33'),
                    test_method_34=Avg('test_method_34'),
                    test_method_35=Avg('test_method_35'),
                    test_method_36=Avg('test_method_36'),
                    test_method_37=Avg('test_method_37'),
                    test_method_38=Avg('test_method_38'),
                    test_method_39=Avg('test_method_39'),
                    test_method_40=Avg('test_method_40'),
                    test_method_41=Avg('test_method_41'),
                    test_method_42=Avg('test_method_42')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'tilt':
                data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'displacement':
                data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'settlement':
                data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'vibration':
                data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            return JsonResponse({
                'code': 200,
                'status': "SUCCESS",
                'result': {'series': series},
            })
        else:
            if sensor_type == 'strain':
                data = models.StrainData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'tilt':
                data = models.TiltData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'displacement':
                data = models.DisplacementData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'settlement':
                data = models.SettlementData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            if sensor_type == 'vibration':
                data = models.VibrationData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                series = []
                sensor_counts = getSensorCounts(sensor_type)
                sensor_names = get_constants(sensor_type)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for index, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{index}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for index, elem in enumerate(request.POST.getlist('method')):
                    series.append(
                        {'color': random_color_code(), 'name': columns[elem], 'data': dynamic_vars[elem]})
            # # Sample data
            # data = {'Date': pd.date_range(start='2022-01-01', periods=10, freq='D'),
            #         'Value': [10, 12, 15, 7, 20, 22, 25, 18, 30, 32]}

            # df = pd.DataFrame(data)
            # df.set_index('Date', inplace=True)

            # # Fit the Holt-Winters Exponential Smoothing model
            # model = ExponentialSmoothing(df['Value'], seasonal='add', seasonal_periods=4)
            # result = model.fit()

            # # Forecast future values
            # forecast_steps = 5
            # forecast = result.forecast(steps=forecast_steps)

            # # Print the forecast values
            # print("Forecasted Values:")
            # print(forecast)
            # exit()
            return JsonResponse({
                'code': 200,
                'status': "SUCCESS",
                'result': {'series': series},
            })
    else:
        return JsonResponse({
            'code': 502,
            'status': "ERROR",
            'message': "There should be ajax method."
        })


def compare(request):
    strain_data = models.StrainData.objects.count()
    if strain_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Strain Data Is Present.')
        return redirect('importExcel')
    tilt_data = models.TiltData.objects.count()
    if tilt_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Tilt Data Is Present.')
        return redirect('importExcel')
    displacement_data = models.DisplacementData.objects.count()
    if displacement_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Displacement Data Is Present.')
        return redirect('importExcel')
    settlement_data = models.SettlementData.objects.count()
    if settlement_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Settlement Data Is Present.')
        return redirect('importExcel')
    vibration_data = models.VibrationData.objects.count()
    if vibration_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Vibration Data Is Present.')
        return redirect('importExcel')
    sensor_types = {
        'keys': list(constants.sensor_types.keys()),
        'values': list(constants.sensor_types.values())
    }
    context.update({'sensor_types': sensor_types})
    return render(request, 'front/compare.html', context)


def compareCombine(request):
    strain_data = models.StrainData.objects.count()
    if strain_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Strain Data Is Present.')
        return redirect('importExcel')
    tilt_data = models.TiltData.objects.count()
    if tilt_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Tilt Data Is Present.')
        return redirect('importExcel')
    displacement_data = models.DisplacementData.objects.count()
    if displacement_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Displacement Data Is Present.')
        return redirect('importExcel')
    settlement_data = models.SettlementData.objects.count()
    if settlement_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Settlement Data Is Present.')
        return redirect('importExcel')
    vibration_data = models.VibrationData.objects.count()
    if vibration_data == 0:
        # Delete all messages by popping them from the list
        try:
            storage = get_messages(request)
            for message in storage:
                message = ''
            storage.used = False
        except:
            pass
        messages.error(request, 'No Vibration Data Is Present.')
        return redirect('importExcel')
    sensor_types = {
        'keys': list(constants.sensor_types.keys()),
        'values': list(constants.sensor_types.values())
    }
    context.update({'sensor_types': sensor_types})
    return render(request, 'front/compare.html', context)


# Getting Min time and Max time by the sensor type
def getCompareTimeDetails(request):
    if request.method == "POST":
        sensor_types = request.POST.getlist('sensor_types[]')
        min_times = []
        max_times = []
        if 'strain' in sensor_types:
            min_times.append((models.StrainData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
            max_times.append((models.StrainData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
        if 'tilt' in sensor_types:
            min_times.append((models.TiltData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
            max_times.append((models.TiltData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
        if 'displacement' in sensor_types:
            min_times.append((models.DisplacementData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
            max_times.append((models.DisplacementData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
        if 'settlement' in sensor_types:
            min_times.append((models.SettlementData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
            max_times.append((models.SettlementData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
        if 'vibration' in sensor_types:
            min_times.append((models.VibrationData.objects.all().order_by(
                'date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
            max_times.append((models.VibrationData.objects.all().order_by(
                '-date_time')[0].date_time).strftime("%Y-%m-%dT%H:%M:%S"))
        # min_date_objects = [date_string for date_string in min_times]
        # Find the maximum datetime object
        min_time = max(min_times)

        # max_date_objects = [date_string for date_string in max_times]
        # Find the minimum datetime object
        max_time = min(max_times)

        return JsonResponse({
            'code': 200,
            'status': "SUCCESS",
            'min_time': removeSeconds(min_time),
            'max_time': removeSeconds(max_time)
        })
    else:
        return JsonResponse({
            'code': 504,
            'status': "ERROR",
            'message': "There should be ajax method."
        })


def getCompareCombineChartData(post_data):
    chart_type = post_data['chart_type']
    from_time = post_data['from_time']
    to_time = post_data['to_time']
    from_miliseconds = int(datetime.fromisoformat(
        from_time).timestamp() * 1000)
    to_miliseconds = int(datetime.fromisoformat(to_time).timestamp() * 1000)
    if from_miliseconds > to_miliseconds:
        return JsonResponse({
            'code': 507,
            'status': "ERROR",
            'message': "From time should not exceeds To time "
        })
    series = []
    for index, elem in enumerate(post_data.getlist('sensor_type')):
        if chart_type == 'max':
            if elem == 'strain':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25'),
                    test_method_26=Max('test_method_26'),
                    test_method_27=Max('test_method_27'),
                    test_method_28=Max('test_method_28'),
                    test_method_29=Max('test_method_29'),
                    test_method_30=Max('test_method_30'),
                    test_method_31=Max('test_method_31'),
                    test_method_32=Max('test_method_32'),
                    test_method_33=Max('test_method_33'),
                    test_method_34=Max('test_method_34'),
                    test_method_35=Max('test_method_35'),
                    test_method_36=Max('test_method_36'),
                    test_method_37=Max('test_method_37'),
                    test_method_38=Max('test_method_38'),
                    test_method_39=Max('test_method_39'),
                    test_method_40=Max('test_method_40'),
                    test_method_41=Max('test_method_41'),
                    test_method_42=Max('test_method_42')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'tilt':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'displacement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'settlement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'vibration':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Max('test_method_1'),
                    test_method_2=Max('test_method_2'),
                    test_method_3=Max('test_method_3'),
                    test_method_4=Max('test_method_4'),
                    test_method_5=Max('test_method_5'),
                    test_method_6=Max('test_method_6'),
                    test_method_7=Max('test_method_7'),
                    test_method_8=Max('test_method_8'),
                    test_method_9=Max('test_method_9'),
                    test_method_10=Max('test_method_10'),
                    test_method_11=Max('test_method_11'),
                    test_method_12=Max('test_method_12'),
                    test_method_13=Max('test_method_13'),
                    test_method_14=Max('test_method_14'),
                    test_method_15=Max('test_method_15'),
                    test_method_16=Max('test_method_16'),
                    test_method_17=Max('test_method_17'),
                    test_method_18=Max('test_method_18'),
                    test_method_19=Max('test_method_19'),
                    test_method_20=Max('test_method_20'),
                    test_method_21=Max('test_method_21'),
                    test_method_22=Max('test_method_22'),
                    test_method_23=Max('test_method_23'),
                    test_method_24=Max('test_method_24'),
                    test_method_25=Max('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
        elif chart_type == 'min':
            if elem == 'strain':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25'),
                    test_method_26=Min('test_method_26'),
                    test_method_27=Min('test_method_27'),
                    test_method_28=Min('test_method_28'),
                    test_method_29=Min('test_method_29'),
                    test_method_30=Min('test_method_30'),
                    test_method_31=Min('test_method_31'),
                    test_method_32=Min('test_method_32'),
                    test_method_33=Min('test_method_33'),
                    test_method_34=Min('test_method_34'),
                    test_method_35=Min('test_method_35'),
                    test_method_36=Min('test_method_36'),
                    test_method_37=Min('test_method_37'),
                    test_method_38=Min('test_method_38'),
                    test_method_39=Min('test_method_39'),
                    test_method_40=Min('test_method_40'),
                    test_method_41=Min('test_method_41'),
                    test_method_42=Min('test_method_42')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'tilt':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'displacement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'settlement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'vibration':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Min('test_method_1'),
                    test_method_2=Min('test_method_2'),
                    test_method_3=Min('test_method_3'),
                    test_method_4=Min('test_method_4'),
                    test_method_5=Min('test_method_5'),
                    test_method_6=Min('test_method_6'),
                    test_method_7=Min('test_method_7'),
                    test_method_8=Min('test_method_8'),
                    test_method_9=Min('test_method_9'),
                    test_method_10=Min('test_method_10'),
                    test_method_11=Min('test_method_11'),
                    test_method_12=Min('test_method_12'),
                    test_method_13=Min('test_method_13'),
                    test_method_14=Min('test_method_14'),
                    test_method_15=Min('test_method_15'),
                    test_method_16=Min('test_method_16'),
                    test_method_17=Min('test_method_17'),
                    test_method_18=Min('test_method_18'),
                    test_method_19=Min('test_method_19'),
                    test_method_20=Min('test_method_20'),
                    test_method_21=Min('test_method_21'),
                    test_method_22=Min('test_method_22'),
                    test_method_23=Min('test_method_23'),
                    test_method_24=Min('test_method_24'),
                    test_method_25=Min('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
        elif chart_type == 'avg':
            if elem == 'strain':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25'),
                    test_method_26=Avg('test_method_26'),
                    test_method_27=Avg('test_method_27'),
                    test_method_28=Avg('test_method_28'),
                    test_method_29=Avg('test_method_29'),
                    test_method_30=Avg('test_method_30'),
                    test_method_31=Avg('test_method_31'),
                    test_method_32=Avg('test_method_32'),
                    test_method_33=Avg('test_method_33'),
                    test_method_34=Avg('test_method_34'),
                    test_method_35=Avg('test_method_35'),
                    test_method_36=Avg('test_method_36'),
                    test_method_37=Avg('test_method_37'),
                    test_method_38=Avg('test_method_38'),
                    test_method_39=Avg('test_method_39'),
                    test_method_40=Avg('test_method_40'),
                    test_method_41=Avg('test_method_41'),
                    test_method_42=Avg('test_method_42')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'tilt':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'displacement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'settlement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'vibration':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                    test_method_1=Avg('test_method_1'),
                    test_method_2=Avg('test_method_2'),
                    test_method_3=Avg('test_method_3'),
                    test_method_4=Avg('test_method_4'),
                    test_method_5=Avg('test_method_5'),
                    test_method_6=Avg('test_method_6'),
                    test_method_7=Avg('test_method_7'),
                    test_method_8=Avg('test_method_8'),
                    test_method_9=Avg('test_method_9'),
                    test_method_10=Avg('test_method_10'),
                    test_method_11=Avg('test_method_11'),
                    test_method_12=Avg('test_method_12'),
                    test_method_13=Avg('test_method_13'),
                    test_method_14=Avg('test_method_14'),
                    test_method_15=Avg('test_method_15'),
                    test_method_16=Avg('test_method_16'),
                    test_method_17=Avg('test_method_17'),
                    test_method_18=Avg('test_method_18'),
                    test_method_19=Avg('test_method_19'),
                    test_method_20=Avg('test_method_20'),
                    test_method_21=Avg('test_method_21'),
                    test_method_22=Avg('test_method_22'),
                    test_method_23=Avg('test_method_23'),
                    test_method_24=Avg('test_method_24'),
                    test_method_25=Avg('test_method_25')
                )
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
        else:
            if elem == 'strain':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.StrainData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'tilt':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.TiltData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'displacement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.DisplacementData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'settlement':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.SettlementData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
            if elem == 'vibration':
                sensor_data = [sensor.split('~~')[1] for sensor in post_data.getlist(
                    'method') if elem + '~~' in sensor]
                data = models.VibrationData.objects.filter(
                    date_time__range=(from_time, to_time)).order_by('date_time')
                sensor_counts = getSensorCounts(elem)
                sensor_names = get_constants(elem)
                columns = {k: v for k, v in sensor_names.items(
                ) if not v.startswith('test_method_')}
                dynamic_vars = {}
                for i, element in enumerate(sensor_counts, start=1):
                    dynamic_vars[f"test_method_{i}"] = []
                for row_data in data:
                    for method in dynamic_vars:
                        dynamic_vars[method].append([int(createMilisecondsByDate(
                            str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                for each in sensor_data:
                    if each in columns.keys():
                        series.append({'color': random_color_code(
                        ), 'name': constants.sensor_types[elem] + "=>" + columns[each], 'data': dynamic_vars[each]})
    random.shuffle(series)
    return series


# Generating compare type chart data by highcharts' recommendation
def getCompareChartData(request):
    if request.method == "POST":
        chart_type = request.POST['chart_type']
        from_time = request.POST['from_time']
        to_time = request.POST['to_time']
        from_miliseconds = int(datetime.fromisoformat(
            from_time).timestamp() * 1000)
        to_miliseconds = int(datetime.fromisoformat(
            to_time).timestamp() * 1000)
        if from_miliseconds > to_miliseconds:
            return JsonResponse({
                'code': 507,
                'status': "ERROR",
                'message': "From time should not exceeds To time "
            })
        all_data = {}
        for index, elem in enumerate(request.POST.getlist('sensor_type')):
            if chart_type == 'max':
                if elem == 'strain':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Max('test_method_1'),
                        test_method_2=Max('test_method_2'),
                        test_method_3=Max('test_method_3'),
                        test_method_4=Max('test_method_4'),
                        test_method_5=Max('test_method_5'),
                        test_method_6=Max('test_method_6'),
                        test_method_7=Max('test_method_7'),
                        test_method_8=Max('test_method_8'),
                        test_method_9=Max('test_method_9'),
                        test_method_10=Max('test_method_10'),
                        test_method_11=Max('test_method_11'),
                        test_method_12=Max('test_method_12'),
                        test_method_13=Max('test_method_13'),
                        test_method_14=Max('test_method_14'),
                        test_method_15=Max('test_method_15'),
                        test_method_16=Max('test_method_16'),
                        test_method_17=Max('test_method_17'),
                        test_method_18=Max('test_method_18'),
                        test_method_19=Max('test_method_19'),
                        test_method_20=Max('test_method_20'),
                        test_method_21=Max('test_method_21'),
                        test_method_22=Max('test_method_22'),
                        test_method_23=Max('test_method_23'),
                        test_method_24=Max('test_method_24'),
                        test_method_25=Max('test_method_25'),
                        test_method_26=Max('test_method_26'),
                        test_method_27=Max('test_method_27'),
                        test_method_28=Max('test_method_28'),
                        test_method_29=Max('test_method_29'),
                        test_method_30=Max('test_method_30'),
                        test_method_31=Max('test_method_31'),
                        test_method_32=Max('test_method_32'),
                        test_method_33=Max('test_method_33'),
                        test_method_34=Max('test_method_34'),
                        test_method_35=Max('test_method_35'),
                        test_method_36=Max('test_method_36'),
                        test_method_37=Max('test_method_37'),
                        test_method_38=Max('test_method_38'),
                        test_method_39=Max('test_method_39'),
                        test_method_40=Max('test_method_40'),
                        test_method_41=Max('test_method_41'),
                        test_method_42=Max('test_method_42')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'tilt':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Max('test_method_1'),
                        test_method_2=Max('test_method_2'),
                        test_method_3=Max('test_method_3'),
                        test_method_4=Max('test_method_4'),
                        test_method_5=Max('test_method_5'),
                        test_method_6=Max('test_method_6'),
                        test_method_7=Max('test_method_7'),
                        test_method_8=Max('test_method_8'),
                        test_method_9=Max('test_method_9'),
                        test_method_10=Max('test_method_10'),
                        test_method_11=Max('test_method_11'),
                        test_method_12=Max('test_method_12'),
                        test_method_13=Max('test_method_13'),
                        test_method_14=Max('test_method_14'),
                        test_method_15=Max('test_method_15'),
                        test_method_16=Max('test_method_16'),
                        test_method_17=Max('test_method_17'),
                        test_method_18=Max('test_method_18'),
                        test_method_19=Max('test_method_19'),
                        test_method_20=Max('test_method_20'),
                        test_method_21=Max('test_method_21'),
                        test_method_22=Max('test_method_22'),
                        test_method_23=Max('test_method_23'),
                        test_method_24=Max('test_method_24'),
                        test_method_25=Max('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'displacement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Max('test_method_1'),
                        test_method_2=Max('test_method_2'),
                        test_method_3=Max('test_method_3'),
                        test_method_4=Max('test_method_4'),
                        test_method_5=Max('test_method_5'),
                        test_method_6=Max('test_method_6'),
                        test_method_7=Max('test_method_7'),
                        test_method_8=Max('test_method_8'),
                        test_method_9=Max('test_method_9'),
                        test_method_10=Max('test_method_10'),
                        test_method_11=Max('test_method_11'),
                        test_method_12=Max('test_method_12'),
                        test_method_13=Max('test_method_13'),
                        test_method_14=Max('test_method_14'),
                        test_method_15=Max('test_method_15'),
                        test_method_16=Max('test_method_16'),
                        test_method_17=Max('test_method_17'),
                        test_method_18=Max('test_method_18'),
                        test_method_19=Max('test_method_19'),
                        test_method_20=Max('test_method_20'),
                        test_method_21=Max('test_method_21'),
                        test_method_22=Max('test_method_22'),
                        test_method_23=Max('test_method_23'),
                        test_method_24=Max('test_method_24'),
                        test_method_25=Max('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'settlement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Max('test_method_1'),
                        test_method_2=Max('test_method_2'),
                        test_method_3=Max('test_method_3'),
                        test_method_4=Max('test_method_4'),
                        test_method_5=Max('test_method_5'),
                        test_method_6=Max('test_method_6'),
                        test_method_7=Max('test_method_7'),
                        test_method_8=Max('test_method_8'),
                        test_method_9=Max('test_method_9'),
                        test_method_10=Max('test_method_10'),
                        test_method_11=Max('test_method_11'),
                        test_method_12=Max('test_method_12'),
                        test_method_13=Max('test_method_13'),
                        test_method_14=Max('test_method_14'),
                        test_method_15=Max('test_method_15'),
                        test_method_16=Max('test_method_16'),
                        test_method_17=Max('test_method_17'),
                        test_method_18=Max('test_method_18'),
                        test_method_19=Max('test_method_19'),
                        test_method_20=Max('test_method_20'),
                        test_method_21=Max('test_method_21'),
                        test_method_22=Max('test_method_22'),
                        test_method_23=Max('test_method_23'),
                        test_method_24=Max('test_method_24'),
                        test_method_25=Max('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'vibration':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Max('test_method_1'),
                        test_method_2=Max('test_method_2'),
                        test_method_3=Max('test_method_3'),
                        test_method_4=Max('test_method_4'),
                        test_method_5=Max('test_method_5'),
                        test_method_6=Max('test_method_6'),
                        test_method_7=Max('test_method_7'),
                        test_method_8=Max('test_method_8'),
                        test_method_9=Max('test_method_9'),
                        test_method_10=Max('test_method_10'),
                        test_method_11=Max('test_method_11'),
                        test_method_12=Max('test_method_12'),
                        test_method_13=Max('test_method_13'),
                        test_method_14=Max('test_method_14'),
                        test_method_15=Max('test_method_15'),
                        test_method_16=Max('test_method_16'),
                        test_method_17=Max('test_method_17'),
                        test_method_18=Max('test_method_18'),
                        test_method_19=Max('test_method_19'),
                        test_method_20=Max('test_method_20'),
                        test_method_21=Max('test_method_21'),
                        test_method_22=Max('test_method_22'),
                        test_method_23=Max('test_method_23'),
                        test_method_24=Max('test_method_24'),
                        test_method_25=Max('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
            elif chart_type == 'min':
                if elem == 'strain':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Min('test_method_1'),
                        test_method_2=Min('test_method_2'),
                        test_method_3=Min('test_method_3'),
                        test_method_4=Min('test_method_4'),
                        test_method_5=Min('test_method_5'),
                        test_method_6=Min('test_method_6'),
                        test_method_7=Min('test_method_7'),
                        test_method_8=Min('test_method_8'),
                        test_method_9=Min('test_method_9'),
                        test_method_10=Min('test_method_10'),
                        test_method_11=Min('test_method_11'),
                        test_method_12=Min('test_method_12'),
                        test_method_13=Min('test_method_13'),
                        test_method_14=Min('test_method_14'),
                        test_method_15=Min('test_method_15'),
                        test_method_16=Min('test_method_16'),
                        test_method_17=Min('test_method_17'),
                        test_method_18=Min('test_method_18'),
                        test_method_19=Min('test_method_19'),
                        test_method_20=Min('test_method_20'),
                        test_method_21=Min('test_method_21'),
                        test_method_22=Min('test_method_22'),
                        test_method_23=Min('test_method_23'),
                        test_method_24=Min('test_method_24'),
                        test_method_25=Min('test_method_25'),
                        test_method_26=Min('test_method_26'),
                        test_method_27=Min('test_method_27'),
                        test_method_28=Min('test_method_28'),
                        test_method_29=Min('test_method_29'),
                        test_method_30=Min('test_method_30'),
                        test_method_31=Min('test_method_31'),
                        test_method_32=Min('test_method_32'),
                        test_method_33=Min('test_method_33'),
                        test_method_34=Min('test_method_34'),
                        test_method_35=Min('test_method_35'),
                        test_method_36=Min('test_method_36'),
                        test_method_37=Min('test_method_37'),
                        test_method_38=Min('test_method_38'),
                        test_method_39=Min('test_method_39'),
                        test_method_40=Min('test_method_40'),
                        test_method_41=Min('test_method_41'),
                        test_method_42=Min('test_method_42')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'tilt':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Min('test_method_1'),
                        test_method_2=Min('test_method_2'),
                        test_method_3=Min('test_method_3'),
                        test_method_4=Min('test_method_4'),
                        test_method_5=Min('test_method_5'),
                        test_method_6=Min('test_method_6'),
                        test_method_7=Min('test_method_7'),
                        test_method_8=Min('test_method_8'),
                        test_method_9=Min('test_method_9'),
                        test_method_10=Min('test_method_10'),
                        test_method_11=Min('test_method_11'),
                        test_method_12=Min('test_method_12'),
                        test_method_13=Min('test_method_13'),
                        test_method_14=Min('test_method_14'),
                        test_method_15=Min('test_method_15'),
                        test_method_16=Min('test_method_16'),
                        test_method_17=Min('test_method_17'),
                        test_method_18=Min('test_method_18'),
                        test_method_19=Min('test_method_19'),
                        test_method_20=Min('test_method_20'),
                        test_method_21=Min('test_method_21'),
                        test_method_22=Min('test_method_22'),
                        test_method_23=Min('test_method_23'),
                        test_method_24=Min('test_method_24'),
                        test_method_25=Min('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'displacement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Min('test_method_1'),
                        test_method_2=Min('test_method_2'),
                        test_method_3=Min('test_method_3'),
                        test_method_4=Min('test_method_4'),
                        test_method_5=Min('test_method_5'),
                        test_method_6=Min('test_method_6'),
                        test_method_7=Min('test_method_7'),
                        test_method_8=Min('test_method_8'),
                        test_method_9=Min('test_method_9'),
                        test_method_10=Min('test_method_10'),
                        test_method_11=Min('test_method_11'),
                        test_method_12=Min('test_method_12'),
                        test_method_13=Min('test_method_13'),
                        test_method_14=Min('test_method_14'),
                        test_method_15=Min('test_method_15'),
                        test_method_16=Min('test_method_16'),
                        test_method_17=Min('test_method_17'),
                        test_method_18=Min('test_method_18'),
                        test_method_19=Min('test_method_19'),
                        test_method_20=Min('test_method_20'),
                        test_method_21=Min('test_method_21'),
                        test_method_22=Min('test_method_22'),
                        test_method_23=Min('test_method_23'),
                        test_method_24=Min('test_method_24'),
                        test_method_25=Min('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'settlement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Min('test_method_1'),
                        test_method_2=Min('test_method_2'),
                        test_method_3=Min('test_method_3'),
                        test_method_4=Min('test_method_4'),
                        test_method_5=Min('test_method_5'),
                        test_method_6=Min('test_method_6'),
                        test_method_7=Min('test_method_7'),
                        test_method_8=Min('test_method_8'),
                        test_method_9=Min('test_method_9'),
                        test_method_10=Min('test_method_10'),
                        test_method_11=Min('test_method_11'),
                        test_method_12=Min('test_method_12'),
                        test_method_13=Min('test_method_13'),
                        test_method_14=Min('test_method_14'),
                        test_method_15=Min('test_method_15'),
                        test_method_16=Min('test_method_16'),
                        test_method_17=Min('test_method_17'),
                        test_method_18=Min('test_method_18'),
                        test_method_19=Min('test_method_19'),
                        test_method_20=Min('test_method_20'),
                        test_method_21=Min('test_method_21'),
                        test_method_22=Min('test_method_22'),
                        test_method_23=Min('test_method_23'),
                        test_method_24=Min('test_method_24'),
                        test_method_25=Min('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'vibration':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Min('test_method_1'),
                        test_method_2=Min('test_method_2'),
                        test_method_3=Min('test_method_3'),
                        test_method_4=Min('test_method_4'),
                        test_method_5=Min('test_method_5'),
                        test_method_6=Min('test_method_6'),
                        test_method_7=Min('test_method_7'),
                        test_method_8=Min('test_method_8'),
                        test_method_9=Min('test_method_9'),
                        test_method_10=Min('test_method_10'),
                        test_method_11=Min('test_method_11'),
                        test_method_12=Min('test_method_12'),
                        test_method_13=Min('test_method_13'),
                        test_method_14=Min('test_method_14'),
                        test_method_15=Min('test_method_15'),
                        test_method_16=Min('test_method_16'),
                        test_method_17=Min('test_method_17'),
                        test_method_18=Min('test_method_18'),
                        test_method_19=Min('test_method_19'),
                        test_method_20=Min('test_method_20'),
                        test_method_21=Min('test_method_21'),
                        test_method_22=Min('test_method_22'),
                        test_method_23=Min('test_method_23'),
                        test_method_24=Min('test_method_24'),
                        test_method_25=Min('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
            elif chart_type == 'avg':
                if elem == 'strain':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.StrainData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Avg('test_method_1'),
                        test_method_2=Avg('test_method_2'),
                        test_method_3=Avg('test_method_3'),
                        test_method_4=Avg('test_method_4'),
                        test_method_5=Avg('test_method_5'),
                        test_method_6=Avg('test_method_6'),
                        test_method_7=Avg('test_method_7'),
                        test_method_8=Avg('test_method_8'),
                        test_method_9=Avg('test_method_9'),
                        test_method_10=Avg('test_method_10'),
                        test_method_11=Avg('test_method_11'),
                        test_method_12=Avg('test_method_12'),
                        test_method_13=Avg('test_method_13'),
                        test_method_14=Avg('test_method_14'),
                        test_method_15=Avg('test_method_15'),
                        test_method_16=Avg('test_method_16'),
                        test_method_17=Avg('test_method_17'),
                        test_method_18=Avg('test_method_18'),
                        test_method_19=Avg('test_method_19'),
                        test_method_20=Avg('test_method_20'),
                        test_method_21=Avg('test_method_21'),
                        test_method_22=Avg('test_method_22'),
                        test_method_23=Avg('test_method_23'),
                        test_method_24=Avg('test_method_24'),
                        test_method_25=Avg('test_method_25'),
                        test_method_26=Avg('test_method_26'),
                        test_method_27=Avg('test_method_27'),
                        test_method_28=Avg('test_method_28'),
                        test_method_29=Avg('test_method_29'),
                        test_method_30=Avg('test_method_30'),
                        test_method_31=Avg('test_method_31'),
                        test_method_32=Avg('test_method_32'),
                        test_method_33=Avg('test_method_33'),
                        test_method_34=Avg('test_method_34'),
                        test_method_35=Avg('test_method_35'),
                        test_method_36=Avg('test_method_36'),
                        test_method_37=Avg('test_method_37'),
                        test_method_38=Avg('test_method_38'),
                        test_method_39=Avg('test_method_39'),
                        test_method_40=Avg('test_method_40'),
                        test_method_41=Avg('test_method_41'),
                        test_method_42=Avg('test_method_42')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'tilt':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.TiltData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Avg('test_method_1'),
                        test_method_2=Avg('test_method_2'),
                        test_method_3=Avg('test_method_3'),
                        test_method_4=Avg('test_method_4'),
                        test_method_5=Avg('test_method_5'),
                        test_method_6=Avg('test_method_6'),
                        test_method_7=Avg('test_method_7'),
                        test_method_8=Avg('test_method_8'),
                        test_method_9=Avg('test_method_9'),
                        test_method_10=Avg('test_method_10'),
                        test_method_11=Avg('test_method_11'),
                        test_method_12=Avg('test_method_12'),
                        test_method_13=Avg('test_method_13'),
                        test_method_14=Avg('test_method_14'),
                        test_method_15=Avg('test_method_15'),
                        test_method_16=Avg('test_method_16'),
                        test_method_17=Avg('test_method_17'),
                        test_method_18=Avg('test_method_18'),
                        test_method_19=Avg('test_method_19'),
                        test_method_20=Avg('test_method_20'),
                        test_method_21=Avg('test_method_21'),
                        test_method_22=Avg('test_method_22'),
                        test_method_23=Avg('test_method_23'),
                        test_method_24=Avg('test_method_24'),
                        test_method_25=Avg('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'displacement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.DisplacementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Avg('test_method_1'),
                        test_method_2=Avg('test_method_2'),
                        test_method_3=Avg('test_method_3'),
                        test_method_4=Avg('test_method_4'),
                        test_method_5=Avg('test_method_5'),
                        test_method_6=Avg('test_method_6'),
                        test_method_7=Avg('test_method_7'),
                        test_method_8=Avg('test_method_8'),
                        test_method_9=Avg('test_method_9'),
                        test_method_10=Avg('test_method_10'),
                        test_method_11=Avg('test_method_11'),
                        test_method_12=Avg('test_method_12'),
                        test_method_13=Avg('test_method_13'),
                        test_method_14=Avg('test_method_14'),
                        test_method_15=Avg('test_method_15'),
                        test_method_16=Avg('test_method_16'),
                        test_method_17=Avg('test_method_17'),
                        test_method_18=Avg('test_method_18'),
                        test_method_19=Avg('test_method_19'),
                        test_method_20=Avg('test_method_20'),
                        test_method_21=Avg('test_method_21'),
                        test_method_22=Avg('test_method_22'),
                        test_method_23=Avg('test_method_23'),
                        test_method_24=Avg('test_method_24'),
                        test_method_25=Avg('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'settlement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.SettlementData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Avg('test_method_1'),
                        test_method_2=Avg('test_method_2'),
                        test_method_3=Avg('test_method_3'),
                        test_method_4=Avg('test_method_4'),
                        test_method_5=Avg('test_method_5'),
                        test_method_6=Avg('test_method_6'),
                        test_method_7=Avg('test_method_7'),
                        test_method_8=Avg('test_method_8'),
                        test_method_9=Avg('test_method_9'),
                        test_method_10=Avg('test_method_10'),
                        test_method_11=Avg('test_method_11'),
                        test_method_12=Avg('test_method_12'),
                        test_method_13=Avg('test_method_13'),
                        test_method_14=Avg('test_method_14'),
                        test_method_15=Avg('test_method_15'),
                        test_method_16=Avg('test_method_16'),
                        test_method_17=Avg('test_method_17'),
                        test_method_18=Avg('test_method_18'),
                        test_method_19=Avg('test_method_19'),
                        test_method_20=Avg('test_method_20'),
                        test_method_21=Avg('test_method_21'),
                        test_method_22=Avg('test_method_22'),
                        test_method_23=Avg('test_method_23'),
                        test_method_24=Avg('test_method_24'),
                        test_method_25=Avg('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'vibration':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.VibrationData.objects.annotate(date=TruncDate(F('date_time'))).values('date').filter(date_time__range=(from_time, to_time)).annotate(
                        test_method_1=Avg('test_method_1'),
                        test_method_2=Avg('test_method_2'),
                        test_method_3=Avg('test_method_3'),
                        test_method_4=Avg('test_method_4'),
                        test_method_5=Avg('test_method_5'),
                        test_method_6=Avg('test_method_6'),
                        test_method_7=Avg('test_method_7'),
                        test_method_8=Avg('test_method_8'),
                        test_method_9=Avg('test_method_9'),
                        test_method_10=Avg('test_method_10'),
                        test_method_11=Avg('test_method_11'),
                        test_method_12=Avg('test_method_12'),
                        test_method_13=Avg('test_method_13'),
                        test_method_14=Avg('test_method_14'),
                        test_method_15=Avg('test_method_15'),
                        test_method_16=Avg('test_method_16'),
                        test_method_17=Avg('test_method_17'),
                        test_method_18=Avg('test_method_18'),
                        test_method_19=Avg('test_method_19'),
                        test_method_20=Avg('test_method_20'),
                        test_method_21=Avg('test_method_21'),
                        test_method_22=Avg('test_method_22'),
                        test_method_23=Avg('test_method_23'),
                        test_method_24=Avg('test_method_24'),
                        test_method_25=Avg('test_method_25')
                    )
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data['date']) + "T00:00:00")), float(row_data[method])])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
            else:
                if elem == 'strain':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.StrainData.objects.filter(
                        date_time__range=(from_time, to_time)).order_by('date_time')
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'tilt':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.TiltData.objects.filter(
                        date_time__range=(from_time, to_time)).order_by('date_time')
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'displacement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.DisplacementData.objects.filter(
                        date_time__range=(from_time, to_time)).order_by('date_time')
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'settlement':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.SettlementData.objects.filter(
                        date_time__range=(from_time, to_time)).order_by('date_time')
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
                if elem == 'vibration':
                    sensor_data = [sensor.split('~~')[1] for sensor in request.POST.getlist(
                        'method') if elem + '~~' in sensor]
                    series = []
                    data = models.VibrationData.objects.filter(
                        date_time__range=(from_time, to_time)).order_by('date_time')
                    sensor_counts = getSensorCounts(elem)
                    sensor_names = get_constants(elem)
                    columns = {k: v for k, v in sensor_names.items(
                    ) if not v.startswith('test_method_')}
                    dynamic_vars = {}
                    for i, element in enumerate(sensor_counts, start=1):
                        dynamic_vars[f"test_method_{i}"] = []
                    for row_data in data:
                        for method in dynamic_vars:
                            dynamic_vars[method].append([int(createMilisecondsByDate(
                                str(row_data.date_time).replace(" ", "T"))), float(getattr(row_data, method))])
                    for each in sensor_data:
                        if each in columns.keys():
                            series.append({'color': random_color_code(
                            ), 'name': columns[each], 'data': dynamic_vars[each]})
                    all_data[elem] = {}
                    all_data[elem]['header'] = "Comparative Data Analysis (" + \
                        constants.sensor_types[elem] + ")"
                    all_data[elem]['data'] = series
        combine_data = getCompareCombineChartData(request.POST)
        return JsonResponse({
            'code': 200,
            'status': "SUCCESS",
            'result': {'all_data': all_data, 'combine_data': combine_data},
        })
    else:
        return JsonResponse({
            'code': 506,
            'status': "ERROR",
            'message': "There should be ajax method."
        })
